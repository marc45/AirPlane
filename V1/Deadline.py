import openpyxl
import subprocess


#打开表格
wb = openpyxl.load_workbook('yuming.xlsx')
# print(wb.sheetnames)
ws = wb.active
#
# # c=ws['B1']
# # print(ws['B1'].value)
# #
# # #coordinate的用法
# # print('cell {} is {}'.format(c.coordinate,c.value))
#
# #行和列
for i in range(1,ws.max_row+1):
    try:
        #遍历Excel表
        DomainName = (ws.cell(row=i,column=2).value).strip()
        # DomainName = (u'\''+ DN+'\b' +'\'')
        print(DomainName)


        # DomainName = 'www.hao123.com.'
        #使用openssl命令
        p1 = subprocess.Popen(['echo'],stdout=subprocess.PIPE,stderr=0)
        p2 = subprocess.Popen(['openssl','s_client','-servername','{}','-connect','{}:443'.format(DomainName,DomainName)],stdin=p1.stdout,stdout=subprocess.PIPE,stderr=0)
        p3 = subprocess.Popen(['openssl','x509','-noout','-dates'], stdin=p2.stdout, stdout=subprocess.PIPE,stderr=0)
        out,err = p3.communicate()
        print(out)
        # print(err)
        # print(out.split('\\n'))

    #将结果写入到表格并保存
        ws.cell(row=i,column=4).value=out
        wb.save('temp4.xlsx')
    except KeyboardInterrupt as base:
        print('ERROR')
        i+=1
        continue

